#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ----------------------------------------------------------------------
# Name:         csv_write_excel.py
# Description:
#
# Author:       m.akei
# Copyright:    (c) 2020 by m.na.akei
# Time-stamp:   <2020-12-28 18:35:51>
# Licence:
#  Copyright (c) 2021 Masaharu N. Akei
#
#  This software is released under the MIT License.
#    http://opensource.org/licenses/mit-license.php
# ----------------------------------------------------------------------
import argparse
import fileinput
import textwrap
import sys

import re

from pathlib import Path

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

from PIL import Image

from io import BytesIO

VERSION = 1.0

IMAGE_EXTENSIONS = ["bmp", "png", "jpg"]
HEIGHT_MARGIN_IMAGE_CELL = 2
WIDTH_MARGIN = 1


def init():
    # argparse --- コマンドラインオプション、引数、サブコマンドのパーサー  Python 3.8.5 ドキュメント https://docs.python.org/ja/3/library/argparse.html
    arg_parser = argparse.ArgumentParser(description="write csv to ms-excel file with images",
                                         formatter_class=argparse.RawDescriptionHelpFormatter,
                                         epilog=textwrap.dedent('''
remark:
  The values, except for numerics,  will be treated as string on excel.

  The datetime values in csv file are converted as string into cells of excel.
  So for treating datetime on excel, the value must be converted by '=DATEVALUE()' and appropriate format on excel.

  Using '--hyperlink', a cell will be with a hyperlink attribute for URL with http/https/file protocol or 
  a value of cell that is available as a path of a file.

example:
  csv_write_excel.py --image_column=file_name --output=images.xlsx images.csv
  csv_write_excel.py --image_column=file_name --output=images.xlsx --hyperlink images.csv

'''))

    arg_parser.add_argument('-v', '--version', action='version', version='%(prog)s {}'.format(VERSION))
    arg_parser.add_argument("--columns",
                            dest="COLUMNS",
                            help="names of columns to write: default=all",
                            type=str,
                            metavar='COLUMN[,COLUMN...]',
                            default=None)
    arg_parser.add_argument("--image_column",
                            dest="IMAGE_COLUMN",
                            help="name of column for path of a image file",
                            type=str,
                            metavar='COLUMN',
                            default="")
    arg_parser.add_argument("--image_output_column",
                            dest="IMAGE_OUTPUT_COLUMN",
                            help="name of column to insert the image, default='Image'",
                            type=str,
                            metavar='COLUMN',
                            default="")
    arg_parser.add_argument("--image_size",
                            dest="IMAGE_SIZE",
                            help="maximum size of image at cell of excel, default='120,120'",
                            type=str,
                            metavar='WIDTH,HEIGHT',
                            default="120,120")
    arg_parser.add_argument("--hyperlink",
                            dest="HYPERLINK",
                            help="enable hyperlink for file path or url",
                            action="store_true",
                            default=False)

    arg_parser.add_argument("--output", dest="OUTPUT", help="path of file to output", type=str, metavar='EXCEL_FILE', required=True)

    arg_parser.add_argument('csv_file', metavar='CSV_FILE', help='csv file to read')
    args = arg_parser.parse_args()
    return args


def trim_output_file(output_file, mkdir=True, overwrite=True):
    """check path of output file

    :param output_file: path of output file
    :param mkdir: if True, do mkdir
    :param overwrite: if False and file existed, then exception will be raised.
    :returns: 
    :rtype: 

    """
    if isinstance(output_file, str):
        p_dir = Path(output_file).parent
        if mkdir and not Path(p_dir).exists():
            Path(p_dir).mkdir(exist_ok=True, parents=True)

        if not overwrite and Path(output_file).exists():
            raise Exception("{} already exists".format(output_file))


def set_image_cell(ws, val, ir, ic, column_width):
    val = XLImage(val)
    column_letter = get_column_letter(ic + 1)
    ws.add_image(val, "{}{}".format(column_letter, ir + 2))
    if column_width[ic] < val.width / 7:
        column_width[ic] = val.width / 7
    ws.row_dimensions[ir + 2].height = pixels_to_points(val.height) + HEIGHT_MARGIN_IMAGE_CELL

    return ws, column_width


def set_string_cell(ws, val, ir, ic, column_width, hyperlink=False):
    if isinstance(val, str) and column_width[ic] < len(val):
        column_width[ic] = len(val)
    ws.cell(row=ir + 1, column=ic + 1).value = val
    ws.cell(row=ir + 1, column=ic + 1).alignment = Alignment(vertical='center')
    if isinstance(val, str):
        if hyperlink and re.search(r"^(http|https|file):/", val) is not None or Path(val).exists():
            ws.cell(row=ir + 1, column=ic + 1).hyperlink = val

    return ws, column_width


def image_resize(img, width, height):
    w, h = img.size
    fact = max(w / width, h / height)

    img = img.resize((int(w / fact), int(h / fact)), resample=Image.BICUBIC)

    return img


def write_to_excel(df, output_file, columns, image_column, image_output_column, image_size, hyperlink=False):

    wb = Workbook()
    # sheet_name = "table_to_excel"
    sheet_name = "Sheet"
    #ws = wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    # python - resize image in openpyxl - Stack Overflow https://stackoverflow.com/questions/59667561/resize-image-in-openpyxl
    if image_output_column not in columns:
        column_width = [0] * (len(columns) + 1)
    else:
        column_width = [0] * len(columns)

    # setup headr row
    for ic in range(len(columns)):
        val = columns[ic]
        ws.cell(row=1, column=ic + 1).value = val
        ws.cell(row=1, column=ic + 1).alignment = Alignment(horizontal='center', vertical='center')
        if isinstance(val, str) and column_width[ic] < len(val):
            column_width[ic] = len(val)

    if image_output_column in columns:
        image_output_column_n = columns.get_loc(image_output_column)
    elif len(image_column) > 0 and len(image_output_column) > 0:
        image_output_column_n = len(columns)
        ws.cell(row=1, column=len(columns) + 1).value = image_output_column
        ws.cell(row=1, column=len(columns) + 1).alignment = Alignment(horizontal='center', vertical='center')
        if isinstance(val, str) and column_width[len(columns)] < len(image_output_column):
            column_width[len(columns)] = len(image_output_column)

    # write each cell
    for ir in range(len(df)):
        for ic in range(len(columns)):
            ic2 = df.columns.get_loc(columns[ic])
            val = df.iat[ir, ic2]
            if image_column == columns[ic]:
                fname = df.iat[ir, ic2]
                # set image cell
                if Path(fname).exists() and Path(fname).suffix[1:] in IMAGE_EXTENSIONS:
                    val = Image.open(fname)
                    val = image_resize(val, image_size[0], image_size[1])
                    fp = BytesIO()
                    val.save(fp, format="png")
                    val = Image.open(fp)
                    ws, column_width = set_image_cell(ws, val, ir, image_output_column_n, column_width)
                ws, column_width = set_string_cell(ws, fname, ir + 1, ic, column_width, hyperlink=hyperlink)
            else:
                ws, column_width = set_string_cell(ws, val, ir + 1, ic, column_width, hyperlink=hyperlink)

    # column width
    if len(image_output_column) > 0 and image_output_column not in columns:
        columns.append(image_output_column)
    for ic in range(len(columns)):
        if column_width[ic] > 0:
            ws.column_dimensions[get_column_letter(ic + 1)].width = column_width[ic] + WIDTH_MARGIN

    wb.save(output_file)


if __name__ == "__main__":
    args = init()
    csv_file = args.csv_file

    columns_s = args.COLUMNS
    image_column = args.IMAGE_COLUMN
    image_output_column = args.IMAGE_OUTPUT_COLUMN
    image_size_s = args.IMAGE_SIZE
    hyperlink_mode = args.HYPERLINK

    xlsx_output = args.OUTPUT

    if csv_file == "-":
        csv_file = sys.stdin

    csv_df = pd.read_csv(csv_file)

    columns = []
    if columns_s is not None:
        columns = re.split(r"\s*,\s*", columns_s)
        for cn in columns:
            if cn not in csv_df.columns:
                print("#warn:csv_write_excel:column '{}' was not found".format(cn), file=sys.stderr)
    else:
        columns = list(csv_df.columns)

    if len(image_column) > 0 and image_column not in csv_df.columns:
        print("??error:csv_write_excel:image column '{}' was not found".format(image_column), file=sys.stderr)
        sys.exit(1)
    if len(image_output_column) > 0 and image_output_column in csv_df.columns:
        print("??warn:csv_write_excel:image column '{}' was found, it will be overwitten.".format(image_output_column), file=sys.stderr)
    if len(image_output_column) > 0 and len(image_column) == 0:
        print(
            "??error:csv_write_excel:image output column '{}' was defined, but image column was not defined.".format(image_output_column),
            file=sys.stderr)
        sys.exit(1)

    if len(image_column) > 0 and len(image_output_column) == 0:
        image_output_column = "Image"

    image_size = list(map(float, re.split(r"\s*,\s*", image_size_s)))

    write_to_excel(csv_df, xlsx_output, columns, image_column, image_output_column, image_size, hyperlink=hyperlink_mode)
